#!/usr/bin/env python3
"""Re-download Regjeringen's kommune number/name xlsx and refresh the bundled JSON.

The JOSM plugin ships a static snapshot under
``core/src/main/resources/no/nvdbincline/core/kommune/kommuner_YYYY-MM-DD.json``.
It does **not** fetch this xlsx at runtime. Re-run this script when Norway's
kommune numbers change (mergers, renumbering).

Requires: Python 3.11+, ``openpyxl``, ``requests``.

Example::

    python tools/refresh_kommune_list.py
    python tools/refresh_kommune_list.py --xlsx /path/to/local.xlsx
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any

try:
    import requests
except ImportError:  # pragma: no cover
    print("error: install requests (pip install requests)", file=sys.stderr)
    sys.exit(1)

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    print("error: install openpyxl (pip install openpyxl)", file=sys.stderr)
    sys.exit(1)

DEFAULT_URL = (
    "https://www.regjeringen.no/contentassets/"
    "832e825c11f145a9b3ce67ef2191662e/"
    "oversikt-over-alle-landets-kommunenummer-og-kommunenavn-per-01.01.2024-excel.xlsx"
)
DEFAULT_EFFECTIVE_DATE = "2024-01-01"
REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUT = (
    REPO_ROOT
    / "core"
    / "src"
    / "main"
    / "resources"
    / "no"
    / "nvdbincline"
    / "core"
    / "kommune"
    / f"kommuner_{DEFAULT_EFFECTIVE_DATE}.json"
)


def _norm(header: str) -> str:
    return header.casefold().strip()


def find_column(headers: list[str], *needles: str) -> int | None:
    """Return first column index whose header contains any needle (casefold)."""
    for i, h in enumerate(headers):
        hn = _norm(h)
        for n in needles:
            if n.casefold() in hn:
                return i
    return None


def find_name_column(headers: list[str]) -> int | None:
    """Prefer 'kommunenavn'; else a 'kommune' column that is not the number."""
    idx = find_column(headers, "kommunenavn")
    if idx is not None:
        return idx
    for i, h in enumerate(headers):
        hn = _norm(h)
        if "kommune" in hn and "nummer" not in hn:
            return i
    return None


def parse_kommune_xlsx(path: Path) -> list[dict[str, Any]]:
    """Parse kommune rows; headers matched dynamically by name, not column index."""
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    if not rows:
        raise ValueError("xlsx is empty")
    headers = [str(c).strip() if c is not None else "" for c in rows[0]]
    num_i = find_column(headers, "kommunenummer")
    name_i = find_name_column(headers)
    if num_i is None or name_i is None:
        raise ValueError(
            "Expected header columns containing 'kommunenummer' and "
            "'kommunenavn' (or 'kommune' for the name). Found headers: "
            + repr(headers)
        )
    out: list[dict[str, Any]] = []
    for row in rows[1:]:
        if row is None or num_i >= len(row) or name_i >= len(row):
            continue
        raw_num = row[num_i]
        raw_name = row[name_i]
        if raw_num is None or raw_name is None:
            continue
        try:
            nummer = int(str(raw_num).strip())
        except ValueError as e:
            raise ValueError(f"non-integer kommunenummer: {raw_num!r}") from e
        navn = str(raw_name).strip()
        if not navn:
            continue
        entry: dict[str, Any] = {"nummer": nummer, "navn": navn}
        # Optional fylke columns if present in a future spreadsheet.
        fylke_num_i = find_column(headers, "fylkesnummer")
        fylke_navn_i = find_column(headers, "fylkesnavn")
        if fylke_num_i is not None and fylke_num_i < len(row) and row[fylke_num_i] is not None:
            try:
                entry["fylkesnummer"] = int(str(row[fylke_num_i]).strip())
            except ValueError:
                pass
        if fylke_navn_i is not None and fylke_navn_i < len(row) and row[fylke_navn_i] is not None:
            fn = str(row[fylke_navn_i]).strip()
            if fn:
                entry["fylkesnavn"] = fn
        out.append(entry)
    if not out:
        raise ValueError("no kommune rows parsed")
    out.sort(key=lambda e: e["nummer"])
    return out


def write_catalog(
    kommuner: list[dict[str, Any]],
    out_path: Path,
    *,
    source: str,
    effective_date: str,
) -> None:
    payload = {
        "source": source,
        "effective_date": effective_date,
        "note": (
            "Bundled snapshot for the JOSM plugin. Do not fetch the live xlsx at "
            "runtime. Refresh with tools/refresh_kommune_list.py when kommune "
            "numbers change. Local-only reference data — not OSM map data."
        ),
        "kommuner": kommuner,
    }
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )


def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--url", default=DEFAULT_URL, help="xlsx download URL")
    p.add_argument("--xlsx", type=Path, help="use local xlsx instead of downloading")
    p.add_argument(
        "--effective-date",
        default=DEFAULT_EFFECTIVE_DATE,
        help="ISO date embedded in metadata and default filename (YYYY-MM-DD)",
    )
    p.add_argument(
        "--out",
        type=Path,
        help="output JSON path (default: core/.../kommuner_<date>.json)",
    )
    args = p.parse_args(argv)

    if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", args.effective_date):
        print("error: --effective-date must be YYYY-MM-DD", file=sys.stderr)
        return 2

    out = args.out or (
        REPO_ROOT
        / "core"
        / "src"
        / "main"
        / "resources"
        / "no"
        / "nvdbincline"
        / "core"
        / "kommune"
        / f"kommuner_{args.effective_date}.json"
    )

    if args.xlsx:
        xlsx = args.xlsx
        source = str(xlsx)
    else:
        xlsx = Path("/tmp") / f"kommuner_{args.effective_date}.xlsx"
        print(f"Downloading {args.url}")
        resp = requests.get(args.url, timeout=60)
        resp.raise_for_status()
        xlsx.write_bytes(resp.content)
        source = args.url

    kommuner = parse_kommune_xlsx(xlsx)
    write_catalog(
        kommuner, out, source=source, effective_date=args.effective_date
    )
    print(f"Wrote {len(kommuner)} kommuner → {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
